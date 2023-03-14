import openpyxl

class Excel_Read:
    Dict = {}


    @staticmethod
    def retrundicvalue(filename,sheetname):

        Dict = {}
        #Code for file name
        book = openpyxl.load_workbook(filename)
        #Code for sheet name
        sheet = book[sheetname]

        #reference code and it is not mantadory reduant code
        cell = sheet.cell(row=1, column=2)
        print(cell.value)
        #get the used row count
        maxrow = sheet.max_row
        print(maxrow)

        # get the used  column
        maxcolumn = sheet.max_column
        print(maxcolumn)

        for i in range(2, maxrow + 1):
            for j in range(1, maxcolumn + 1):
                Dict[(sheet.cell(row=1, column=j).value) + str(i - 1)] = sheet.cell(row=i, column=j).value

        print(Dict)

        return Dict



    @staticmethod
    def retrunmaxrow(filename,sheetname):
        book = openpyxl.load_workbook(filename)
        sheet = book[sheetname]

        #get the used row count
        maxrow = sheet.max_row
        print(maxrow)

        return maxrow

    @staticmethod
    def retrunmaxcolumn(filename,sheetname):
        book = openpyxl.load_workbook(filename)
        sheet = book[sheetname]

        #get the used row count
        maxcolumn = sheet.max_column
        print(maxcolumn)

        return maxcolumn